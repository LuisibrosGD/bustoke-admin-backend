import io
from typing import Any


async def export_excel(data: list[dict[str, Any]], sheet_name: str = "Reporte") -> bytes:
    """Genera un archivo Excel en memoria y devuelve los bytes."""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    if not data:
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    headers = list(data[0].keys())
    ws.append(headers)
    for row in data:
        ws.append([row.get(h) for h in headers])

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def read_excel_rows(file_bytes: bytes) -> list[dict[str, Any]]:
    """Lee la primera hoja de un .xlsx: primera fila = headers, cada fila
    siguiente se devuelve como dict {header: valor}. Ignora filas vacías."""
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    headers = [str(h).strip() if h else "" for h in next(rows_iter, [])]
    rows = []
    for raw in rows_iter:
        if all(v is None for v in raw):
            continue
        rows.append({headers[i]: raw[i] for i in range(len(headers)) if i < len(raw)})
    return rows
